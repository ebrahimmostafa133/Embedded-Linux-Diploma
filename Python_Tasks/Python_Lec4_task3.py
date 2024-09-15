#---------  lec 4 ----------
# author: Ebrahim Mostafa

#------------------------------------------------ Task 3 ------------------------------------------------
'''
Write a python code that manage a database for employees.
Each employee has a unique ID and has the following data:
● Name, Job and Salary. The system shall allow:
● 1-Add new employee
● 2-Print employee data
● 3-remove employee from the system
● 4- update

'''
import openpyxl

# Function to initialize the Excel sheet
def initialize_excel(file_path):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Add headers to the worksheet
    ws.append(["ID", "Name", "Job", "Salary"])

    # Save the workbook to the specified file path
    wb.save(file_path)
    print(f"Initialized new Excel sheet at {file_path}")

# Function to load or create the Excel sheet
def load_or_create_excel(file_path):
    try:
        # Try to load the existing Excel file
        wb = openpyxl.load_workbook(file_path)
        print(f"Loaded existing Excel sheet from {file_path}")
    except FileNotFoundError:
        # If the file does not exist, initialize a new one
        initialize_excel(file_path)
        wb = openpyxl.load_workbook(file_path)
    return wb

# Function to add a new employee
def add_employee(wb, file_path, emp_id, name, job, salary):
    ws = wb.active
    # Check if employee ID already exists
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == emp_id:
            print(f"Employee ID {emp_id} already exists.")
            return

    # Append the new employee data to the worksheet
    ws.append([emp_id, name, job, salary])
    wb.save(file_path)
    print(f"Employee {name} added successfully!")

# Function to print employee data
def print_employee_data(wb):
    ws = wb.active
    # Print the header row
    print(f"{ws['A1'].value}\t{ws['B1'].value}\t{ws['C1'].value}\t{ws['D1'].value}")

    # Print each employee's data
    for row in ws.iter_rows(min_row=2, values_only=True):
        print(f"{row[0]}\t{row[1]}\t{row[2]}\t{row[3]}")

# Function to remove an employee
def remove_employee(wb, file_path, emp_id):
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == emp_id:
            ws.delete_rows(row[0].row, 1)
            wb.save(file_path)
            print(f"Employee with ID {emp_id} removed successfully!")
            return
    print(f"No employee found with ID {emp_id}.")

# Function to update an employee
def update_employee(wb, file_path, emp_id, name=None, job=None, salary=None):
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == emp_id:
            if name:
                row[1].value = name
            if job:
                row[2].value = job
            if salary:
                row[3].value = salary
            wb.save(file_path)
            print(f"Employee with ID {emp_id} updated successfully!")
            return
    print(f"No employee found with ID {emp_id}.")

# Main loop for user interaction
def main():
    file_path = "employees.xlsx"
    wb = load_or_create_excel(file_path)

    while True:
        print("\nEmployee Management System")
        print("1 - Add New Employee")
        print("2 - Print Employee Data")
        print("3 - Remove Employee")
        print("4 - Update Employee")
        print("5 - Exit")

        choice = input("Enter your choice: ")

        if choice == '1':
            emp_id = input("Enter employee ID: ")
            name = input("Enter employee name: ")
            job = input("Enter employee job: ")
            salary = float(input("Enter employee salary: "))
            add_employee(wb, file_path, emp_id, name, job, salary)

        elif choice == '2':
            print_employee_data(wb)

        elif choice == '3':
            emp_id = input("Enter employee ID to remove: ")
            remove_employee(wb, file_path, emp_id)

        elif choice == '4':
            emp_id = input("Enter employee ID to update: ")
            print("Enter new details (leave blank to keep unchanged):")
            name = input("New name: ")
            job = input("New job: ")
            salary_input = input("New salary: ")

            # Convert salary to float if provided
            salary = float(salary_input) if salary_input else None

            update_employee(wb, file_path, emp_id, name=name if name else None, job=job if job else None, salary=salary)

        elif choice == '5':
            print("Exiting Employee Management System.")
            break

        else:
            print("Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
